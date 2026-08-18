"""French CIQUAL 2025 adapter (ANSES, FR edition — the EN 2020 file is legacy
`.xls` that openpyxl cannot read).

Cell grammar: comma decimals ("2,89"), censored "< 0,5" (kept as the upper
bound with quality `censored`), "traces" -> 0, "-" -> not reported. CIQUAL
does not publish per-value origins in this export, so rows are `compiled`;
the sweep found raw-meat entries can be verbatim USDA copies — the echo
screener is the guard, always extract with it on.

Vitamin K: K1 and K2 are separate columns; the platform's 1185 row follows
the total-K doctrine, so K1+K2 are summed here with the forms noted.
"""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Optional

from intake.model import Q_CENSORED, Q_COMPILED, Q_COMPUTED, Q_TRACE, SourceValue
from intake.units import to_fediaf

LABEL = "CIQUAL"
_XLSX = Path(__file__).resolve().parents[2] / "data" / "ciqual_fr" / "Table_Ciqual_2025_FR.xlsx"

# column index -> (nutrient_id, unit, short label)
COL_MAP: dict[int, tuple[int, str, str]] = {
    12: (1008, "kcal", "Energie NxJones"),
    13: (1051, "g", "Eau"),
    15: (1003, "g", "Protéines Nx6.25"),
    16: (1005, "g", "Glucides"),
    17: (1004, "g", "Lipides"),
    26: (1079, "g", "Fibres"),
    28: (1007, "g", "Cendres"),
    43: (1269, "g", "AG 18:2 LA"), 44: (1270, "g", "AG 18:3 ALA"),
    45: (1271, "g", "AG 20:4 ARA"), 46: (1278, "g", "AG 20:5 EPA"),
    47: (1272, "g", "AG 22:6 DHA"),
    50: (1087, "mg", "Calcium"), 51: (1088, "mg", "Chlorure"),
    52: (1098, "mg", "Cuivre"), 53: (1089, "mg", "Fer"),
    54: (1100, "µg", "Iode"), 55: (1090, "mg", "Magnésium"),
    56: (1101, "mg", "Manganèse"), 57: (1091, "mg", "Phosphore"),
    58: (1092, "mg", "Potassium"), 59: (1103, "µg", "Sélénium"),
    60: (1093, "mg", "Sodium"), 61: (1095, "mg", "Zinc"),
    63: (1106, "µg", "Rétinol"), 65: (1110, "µg", "Vitamine D"),
    68: (1109, "mg", "Alpha-tocophérol"),
    73: (1165, "mg", "B1"), 74: (1166, "mg", "B2"), 75: (1167, "mg", "B3 niacine"),
    76: (1170, "mg", "B5"), 77: (1175, "mg", "B6"),
    79: (1177, "µg", "Folates totaux"), 82: (1178, "µg", "B12"),
}
_K1_COL, _K2_COL = 70, 71
_COMPUTED_COLS = frozenset({12})  # energy is always calculated


def _parse(raw) -> Optional[tuple[float, str]]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw), Q_COMPILED
    text = str(raw).strip().replace(",", ".")
    if text in ("-", ""):
        return None
    if text.lower() == "traces":
        return 0.0, Q_TRACE
    if text.startswith("<"):
        try:
            return float(text[1:].strip()), Q_CENSORED
        except ValueError:
            return None
    try:
        return float(text), Q_COMPILED
    except ValueError:
        return None


@lru_cache(maxsize=1)
def _rows() -> dict[int, tuple]:
    from openpyxl import load_workbook

    wb = load_workbook(_XLSX, read_only=True, data_only=True)
    ws = wb["composition nutritionnelle"]
    it = ws.iter_rows(values_only=True)
    next(it)
    out = {}
    for r in it:
        if r[6] is not None:
            out[int(r[6])] = tuple(r)
    wb.close()
    return out


def extract(key: int, note: str = "") -> list[SourceValue]:
    row = _rows().get(int(key))
    if row is None:
        raise KeyError(f"CIQUAL alim_code {key} not found")
    food = f"{key} {row[7]}"
    out: list[SourceValue] = []
    for i, (nid, unit, label) in COL_MAP.items():
        parsed = _parse(row[i])
        if parsed is None:
            continue
        value, quality = parsed
        if i in _COMPUTED_COLS:
            quality = Q_COMPUTED
        out.append(SourceValue(
            source=LABEL, source_food=food, nutrient_id=nid,
            value=to_fediaf(nid, value, unit), quality=quality, note=label,
        ))
    # vitamin K: total-K doctrine (K1 + K2 when reported)
    k1, k2 = _parse(row[_K1_COL]), _parse(row[_K2_COL])
    if k1 is not None or k2 is not None:
        total = (k1[0] if k1 else 0.0) + (k2[0] if k2 else 0.0)
        quality = Q_CENSORED if (k1 and k1[1] == Q_CENSORED) or (k2 and k2[1] == Q_CENSORED) else Q_COMPILED
        forms = []
        forms.append(f"K1={k1[0]:g}" if k1 else "K1 n/a")
        forms.append(f"K2={k2[0]:g}" if k2 else "K2 n/a")
        out.append(SourceValue(
            source=LABEL, source_food=food, nutrient_id=1185,
            value=to_fediaf(1185, total, "µg"), quality=quality,
            note="Vitamine K1+K2 total-K; " + ", ".join(forms),
        ))
    return out


def search(query: str) -> list[tuple[int, str]]:
    q = query.lower()
    return [
        (code, str(r[7])) for code, r in sorted(_rows().items())
        if q in str(r[7] or "").lower()
    ]
