"""USDA/FDA/ODS-NIH Iodine Database Release 4 adapter (per-100 g workbook).

The only iodine source with real sampling statistics (n, SD, min, max) keyed
to SR/Foundation NDB numbers. USDA-affiliated — analytical values absent from
FDC, so it is treated as measured evidence but NOT as an independent foreign
table for echo purposes. Group-header rows are interleaved in the sheet and
are skipped by requiring a numeric value column.
"""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path

from intake.model import Q_ANALYSED, SourceValue
from intake.units import check_header_alignment, parse_float

LABEL = "IodineDB"
_XLSX = (Path(__file__).resolve().parents[2]
         / "data" / "usda_iodine" / "Iodine Database_Release 4_Per 100g.xlsx")
IODINE_ID = 1100

# source-level lineage: every value from this adapter is USDA-affiliated
# (consumed centrally by extract.run — never independent confirmation)
USDA_LINEAGE = True

_f = parse_float

# header row 2 (0-indexed 1): the positional map this adapter relies on
_HEADER_LABELS = {3: "Description", 4: "n", 5: "Iodine", 6: "SD",
                  7: "Min", 8: "Max"}


@lru_cache(maxsize=1)
def _rows() -> list[tuple]:
    from openpyxl import load_workbook

    wb = load_workbook(_XLSX, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    all_rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    check_header_alignment(all_rows[1], _HEADER_LABELS, LABEL)
    return all_rows[2:]


def extract(key: str, note: str = "") -> list[SourceValue]:
    """`key` is the 5-digit NDB number (column B), e.g. '05096'."""
    for r in _rows():
        if str(r[1] or "").strip() == str(key) and _f(r[5]) is not None:
            n = _f(r[4])
            sd = _f(r[6])
            return [SourceValue(
                source=LABEL, source_food=f"NDB {key} {r[3]}",
                nutrient_id=IODINE_ID, value=float(r[5]),
                n=int(n) if n else None, vmin=_f(r[7]), vmax=_f(r[8]),
                quality=Q_ANALYSED,
                note=f"Iodine DB R4; SD={sd if sd is not None else 'n/a'}",
                usda_lineage=True,  # USDA/FDA/ODS program: evidence, never
                                    # independent confirmation of USDA
            )]
    raise KeyError(f"Iodine DB NDB number {key!r} not found")


def search(query: str) -> list[tuple[str, str]]:
    q = query.lower()
    return [
        (str(r[1]), str(r[3])) for r in _rows()
        if _f(r[5]) is not None and q in str(r[3] or "").lower() and str(r[1] or "").strip()
    ]
