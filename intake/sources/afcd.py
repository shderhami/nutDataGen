"""Australian AFCD Release 3 adapter (FSANZ).

`Nutrient_profiles` sheet `All solids & liquids per 100 g`, header row 3.
Fatty acids exist as %T and as mass columns — ONLY the mass columns are read
(the %T trap). Amino acids exist per g N and as mass columns — only mass.
The food-level `Derivation` (col 2) distinguishes `Analysed` from `Recipe`/
borrowed rows.

`Food_Details -> Sampling Details` names which nutrients were imputed or
borrowed (AFCD tells you!). Sentences containing impute/borrow markers are
scanned for nutrient names and those values are downgraded to `borrowed`.
"""
from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path

from intake.model import Q_ANALYSED, Q_BORROWED, Q_COMPUTED, SourceValue
from intake.units import check_header_alignment, parse_float, to_fediaf

LABEL = "AFCD"
_DIR = Path(__file__).resolve().parents[2] / "data" / "afcd_au"
_PROFILES = _DIR / "AFCD_R3_Nutrient_profiles.xlsx"
_DETAILS = _DIR / "AFCD_R3_Food_Details.xlsx"

# column index -> (nutrient_id, unit, label). Units mirror the header text.
COL_MAP: dict[int, tuple[int, str, str]] = {
    5: (1008, "kJ", "Energy without fibre, equated"),
    6: (1051, "g", "Moisture"),
    7: (1003, "g", "Protein"),
    9: (1004, "g", "Fat, total"),
    10: (1007, "g", "Ash"),
    11: (1079, "g", "Total dietary fibre"),
    38: (1005, "g", "Available carbohydrate w/o sugar alcohols"),
    55: (1087, "mg", "Calcium"), 57: (1088, "mg", "Chloride"),
    59: (1098, "mg", "Copper"), 61: (1100, "µg", "Iodine"),
    62: (1089, "mg", "Iron"), 64: (1090, "mg", "Magnesium"),
    65: (1101, "mg", "Manganese"), 69: (1091, "mg", "Phosphorus"),
    70: (1092, "mg", "Potassium"), 71: (1103, "µg", "Selenium"),
    72: (1093, "mg", "Sodium"), 75: (1095, "mg", "Zinc"),
    76: (1106, "µg", "Retinol (preformed)"),
    85: (1165, "mg", "Thiamin"), 86: (1166, "mg", "Riboflavin"),
    87: (1167, "mg", "Niacin (B3)"),  # preformed niacin, matches USDA 1167
    90: (1170, "mg", "Pantothenic acid"), 91: (1175, "mg", "Pyridoxine B6"),
    92: (1176, "µg", "Biotin"), 93: (1178, "µg", "Cobalamin B12"),
    94: (1177, "µg", "Folate, natural"),
    103: (1110, "µg", "Vitamin D3 equivalents"),
    104: (1109, "mg", "Alpha tocopherol"),
    208: (1269, "g", "C18:2w6 LA"), 209: (1270, "g", "C18:3w3 ALA"),
    221: (1271, "mg", "C20:4w6 ARA"), 222: (1278, "mg", "C20:5w3 EPA"),
    224: (1280, "mg", "C22:5w3 DPA"), 229: (1272, "mg", "C22:6w3 DHA"),
    230: (1293, "g", "Total PUFA, equated"),
    255: (1220, "mg", "Arginine"), 257: (1216, "mg", "Cystine plus cysteine"),
    260: (1221, "mg", "Histidine"), 261: (1212, "mg", "Isoleucine"),
    262: (1213, "mg", "Leucine"), 263: (1214, "mg", "Lysine"),
    264: (1215, "mg", "Methionine"), 265: (1217, "mg", "Phenylalanine"),
    268: (1211, "mg", "Threonine"), 269: (1218, "mg", "Tyrosine"),
    270: (1210, "mg", "Tryptophan"), 271: (1219, "mg", "Valine"),
}
_COMPUTED_COLS = frozenset({5})

# nutrient-name keywords for the sampling-details borrow scan
_BORROW_KEYWORDS: dict[str, tuple[int, ...]] = {
    "vitamin d": (1110,), "vitamin e": (1109,), "tocopherol": (1109,),
    "biotin": (1176,), "folate": (1177,), "vitamin b12": (1178,),
    "iodine": (1100,), "selenium": (1103,), "choline": (1180,),
    "fatty acid": (1269, 1270, 1271, 1272, 1278, 1280, 1293),
    "amino acid": (1210, 1211, 1212, 1213, 1214, 1215, 1216, 1217, 1218,
                   1219, 1220, 1221),
    "vitamin": (),  # generic mention: kept for the note only
}


_f = parse_float


@lru_cache(maxsize=1)
def _profiles() -> dict[str, tuple]:
    from openpyxl import load_workbook

    wb = load_workbook(_PROFILES, read_only=True, data_only=True)
    ws = wb["All solids & liquids per 100 g"]
    it = ws.iter_rows(values_only=True)
    next(it), next(it)
    header = next(it)
    check_header_alignment(
        header, {i: label for i, (_nid, _unit, label) in COL_MAP.items()}, LABEL)
    out = {str(r[0]): tuple(r) for r in it if r[0] is not None}
    wb.close()
    return out


@lru_cache(maxsize=1)
def _details() -> dict[str, dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(_DETAILS, read_only=True, data_only=True)
    ws = wb["Food details"]
    rows = ws.iter_rows(values_only=True)
    hdr = None
    for r in rows:
        if any("Public Food Key" in str(c) for c in r if c):
            hdr = [str(h or "") for h in r]
            break
    out: dict[str, dict[str, str]] = {}
    if hdr:
        for r in rows:
            key = str(r[0] or "")
            if key:
                out[key] = {h: str(v) for h, v in zip(hdr, r) if v not in (None, "")}
    wb.close()
    return out


def _borrowed_ids(sampling: str) -> tuple[set[int], list[str]]:
    """Nutrient ids named in impute/borrow sentences, plus those sentences."""
    flagged: set[int] = set()
    sentences: list[str] = []
    for sentence in re.split(r"(?<=[.;])\s+", sampling):
        lowered = sentence.lower()
        if not any(m in lowered for m in ("imput", "borrow", "derived from usda",
                                          "taken from", "based on data for")):
            continue
        sentences.append(sentence.strip())
        for kw, ids in _BORROW_KEYWORDS.items():
            if kw in lowered:
                flagged.update(ids)
    return flagged, sentences


def extract(key: str, note: str = "") -> list[SourceValue]:
    row = _profiles().get(str(key))
    if row is None:
        raise KeyError(f"AFCD key {key!r} not found")
    food = f"{key} {row[3]}"
    derivation = str(row[2] or "")
    base_quality = Q_ANALYSED if derivation.lower() == "analysed" else Q_COMPUTED
    detail = _details().get(str(key), {})
    sampling = detail.get("Sampling Details", "")
    flagged, sentences = _borrowed_ids(sampling)

    out: list[SourceValue] = []
    for i, (nid, unit, label) in COL_MAP.items():
        val = _f(row[i])
        if val is None:
            continue
        quality = base_quality
        item_note = f"{label}; food-level derivation: {derivation}"
        if i in _COMPUTED_COLS:
            quality = Q_COMPUTED
        if nid in flagged:
            quality = Q_BORROWED
            item_note += f"; sampling: {sentences[0][:120]}" if sentences else ""
        out.append(SourceValue(
            source=LABEL, source_food=food, nutrient_id=nid,
            value=to_fediaf(nid, val, unit), quality=quality, note=item_note,
        ))
    return out


def search(query: str) -> list[tuple[str, str]]:
    q = query.lower()
    return [
        (key, str(r[3])) for key, r in sorted(_profiles().items())
        if q in str(r[3] or "").lower()
    ]
