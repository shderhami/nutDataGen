"""Danish FCDB 6.1 adapter (DTU; ex-Frida).

Reads `Data_Normalised` (long format: per-value Source ids, Min/Max/Median,
NumberOfDeterminations) — the sheet the sweep found to be gold when the Source
decodes to Danish analytical work. Source ids are decoded via the `Source`
sheet; borrowed rows (USDA, McCance, Fineli, Souci, ...) are flagged, they are
NOT independent confirmation. Units come from the `Parameter` sheet.
"""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path

from intake.model import (
    Q_ANALYSED,
    Q_BORROWED,
    Q_COMPUTED,
    Q_ESTIMATED,
    Q_UNKNOWN,
    SourceValue,
)
from intake.units import parse_float, parse_per100g_unit, to_fediaf

LABEL = "FCDB"
_XLSX = Path(__file__).resolve().parents[2] / "data" / "fcdb_dk" / "FCDB_6.1_Dataset.xlsx"

# ParameterName -> FEDIAF nutrient id (definition-comparable only).
# NOTE "Vitamin D" (total) is the definitional match for 1110 (D2+D3); the CV
# pipeline's cv_intl.PARAM_MAP reads the "Vitamin D3" row for its dispersion
# observations — a deliberate, documented split, do not "fix" one to the other.
PARAM_MAP: dict[str, int] = {
    "Energy (kcal)": 1008,
    "Protein": 1003, "Fat": 1004, "Ash": 1007, "Water": 1051,
    "Carbohydrate by difference": 1005, "Dietary fibre": 1079,
    "Retinol": 1106, "Vitamin D": 1110, "alpha-Tocopherol": 1109, "Vitamin K": 1185,
    "Thiamin (Vitamin B1)": 1165, "Riboflavin (Vitamin B2)": 1166, "Niacin": 1167,
    "Vitamin B6": 1175, "Vitamin B12": 1178, "Pantothenic acid": 1170,
    "Folate": 1177, "Biotin": 1176, "Choline": 1180,
    "Sodium": 1093, "Potassium": 1092, "Calcium": 1087, "Magnesium": 1090,
    "Iron": 1089, "Copper": 1098, "Zinc": 1095, "Manganese": 1101,
    "Selenium": 1103, "Phosphorus": 1091, "Iodine": 1100, "Chloride": 1088,
    "C18:2,n-6": 1269, "C18:3,n-3": 1270, "C20:4,n-6": 1271,
    "C20:5,n-3": 1278, "C22:5,n-3": 1280, "C22:6,n-3": 1272,
    "Sum polyunsaturated fatty acids": 1293,
    "Isoleucine": 1212, "Leucine": 1213, "Lysine": 1214, "Methionine": 1215,
    "Cystine": 1216, "Phenylalanine": 1217, "Tyrosine": 1218, "Threonine": 1211,
    "Tryptophan": 1210, "Valine": 1219, "Arginine": 1220, "Histidine": 1221,
    "Taurine": 1234,
}

_BORROWED_MARKERS = (
    "usda", "fooddata central", "food data central", "mccance", "widdowson",
    "fineli", "souci", "swedish", "livsmedelsverket",
    "food composition table", "nubel", "danish food composition databank ed. 7",
)
_COMPUTED_MARKERS = ("calculat", "beregn", "recipe")
_ESTIMATED_MARKERS = ("estimated value", "anslået", "skøn")


_f = parse_float


@lru_cache(maxsize=1)
def _meta() -> tuple[dict[str, str], dict[int, str], dict[int, str]]:
    """(param->unit, source id->citation, food id->name) — metadata sheets only."""
    from openpyxl import load_workbook

    wb = load_workbook(_XLSX, read_only=True, data_only=True)

    ws = wb["Parameter"]
    it = ws.iter_rows(values_only=True)
    hdr = {h: i for i, h in enumerate(next(it))}
    param_units = {
        str(r[hdr["ParameterName"]]): str(r[hdr["Unit"]])
        for r in it if r[hdr["ParameterName"]] is not None
    }

    ws = wb["Source"]
    it = ws.iter_rows(values_only=True)
    hdr = {h: i for i, h in enumerate(next(it))}
    sources = {}
    for r in it:
        sid = _f(r[hdr["SourceID"]])
        if sid is None:
            continue
        title = str(r[hdr["TitleEnglish"]])
        if title in ("NULL", "None", ""):
            title = str(r[hdr["TitleOriginal"]])
        year = str(r[hdr["Year"]])
        sources[int(sid)] = f"{title} ({year})" if year not in ("NULL", "None", "") else title

    ws = wb["Food"]
    it = ws.iter_rows(values_only=True)
    hdr = {h: i for i, h in enumerate(next(it))}
    foods = {
        int(r[hdr["FoodID"]]): str(r[hdr["FoodName"]])
        for r in it if _f(r[hdr["FoodID"]]) is not None
    }
    wb.close()
    return param_units, sources, foods


@lru_cache(maxsize=1)
def _data_by_food() -> dict[int, list[tuple]]:
    """Data_Normalised rows for mapped parameters only, grouped by FoodID.

    One streaming pass; keeps ~50 rows/food instead of the full 144k-row
    sheet (which cost ~46 MB resident and a second full copy)."""
    from openpyxl import load_workbook

    wb = load_workbook(_XLSX, read_only=True, data_only=True)
    ws = wb["Data_Normalised"]
    it = ws.iter_rows(values_only=True)
    idx = {h: i for i, h in enumerate(next(it))}
    out: dict[int, list[tuple]] = {}
    for r in it:
        param = str(r[idx["ParameterName"]])
        if param not in PARAM_MAP:
            continue
        # tolerate float/str-typed ids like _meta() does — a re-export that
        # changes the cell type must not silently empty the extraction
        food_id_raw = _f(r[idx["FoodID"]])
        if food_id_raw is None:
            continue
        food_id = int(food_id_raw)
        out.setdefault(food_id, []).append((
            param, _f(r[idx["ResVal"]]), _f(r[idx["Min"]]), _f(r[idx["Max"]]),
            r[idx["NumberOfDeterminations"]], str(r[idx["Source"]]),
        ))
    wb.close()
    return out


def _classify(source_ids: str, sources: dict[int, str]) -> tuple[str, str]:
    """(quality, note) from the row's comma-separated Source ids."""
    texts = []
    for tok in source_ids.replace("NULL", "").split(","):
        tok = tok.strip()
        if tok.isdigit():
            texts.append(sources.get(int(tok), f"src {tok}?"))
    joined = " | ".join(texts)
    lowered = joined.lower()
    if any(m in lowered for m in _ESTIMATED_MARKERS):
        quality = Q_ESTIMATED
    elif any(m in lowered for m in _COMPUTED_MARKERS):
        quality = Q_COMPUTED
    elif any(m in lowered for m in _BORROWED_MARKERS):
        quality = Q_BORROWED
    elif texts:
        quality = Q_ANALYSED
    else:
        quality = Q_UNKNOWN
    return quality, f"src [{source_ids}] {joined[:160]}" if texts else ""


def extract(key: int, note: str = "") -> list[SourceValue]:
    param_units, sources, foods = _meta()
    if int(key) not in foods:
        raise KeyError(f"FCDB FoodID {key} not found")
    food_name = foods[int(key)]
    out: list[SourceValue] = []
    for param, val, vmin, vmax, n, src in _data_by_food().get(int(key), []):
        if val is None:
            continue
        nid = PARAM_MAP[param]
        unit = parse_per100g_unit(param_units[param])
        quality, src_note = _classify(src, sources)
        n_val = _f(n)
        n_int = int(n_val) if n_val is not None and n_val > 0 else None
        scale = to_fediaf(nid, 1.0, unit)
        out.append(SourceValue(
            source=LABEL, source_food=f"{key} {food_name}", nutrient_id=nid,
            value=val * scale, n=n_int,
            vmin=vmin * scale if vmin is not None else None,
            vmax=vmax * scale if vmax is not None else None,
            quality=quality,
            note=f"{param}; {src_note}" if src_note else param,
        ))
    return out


def search(query: str) -> list[tuple[int, str]]:
    _, _, foods = _meta()
    q = query.lower()
    return [(fid, name) for fid, name in sorted(foods.items()) if q in name.lower()]
