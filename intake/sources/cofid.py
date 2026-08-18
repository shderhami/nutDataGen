"""UK CoFID 2021 adapter (McCance & Widdowson's integrated dataset).

A compilation — quality is `compiled` and each value's note carries the
sheet's `Main data references` so the operator can cite the underlying survey
(sweep rule: never cite "CoFID" as the source). Cell grammar: `N` = no
reliable information (skipped), `Tr` = trace -> 0. Vitamin K is K1 only.
Fatty acids are read from `1.12 (PUFA per 100gFood)` — never the /100g-FA
sheets. CoFID publishes no ash and no amino acids.
"""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Optional

from intake.model import Q_COMPILED, Q_TRACE, SourceValue
from intake.units import to_fediaf

LABEL = "CoFID"
_XLSX = (Path(__file__).resolve().parents[2]
         / "data" / "McCance_Widdowsons_Composition_of_Foods_Integrated_Dataset_2021..xlsx")

# sheet -> {column index: (nutrient_id, unit, label)}
SHEET_MAP: dict[str, dict[int, tuple[int, str, str]]] = {
    "1.3 Proximates": {
        7: (1051, "g", "Water"), 9: (1003, "g", "Protein"), 10: (1004, "g", "Fat"),
        11: (1005, "g", "Carbohydrate"), 12: (1008, "kcal", "Energy"),
        25: (1079, "g", "AOAC fibre"),
        39: (1293, "g", "Poly FA /100g food"),
    },
    "1.4 Inorganics": {
        7: (1093, "mg", "Sodium"), 8: (1092, "mg", "Potassium"),
        9: (1087, "mg", "Calcium"), 10: (1090, "mg", "Magnesium"),
        11: (1091, "mg", "Phosphorus"), 12: (1089, "mg", "Iron"),
        13: (1098, "mg", "Copper"), 14: (1095, "mg", "Zinc"),
        15: (1088, "mg", "Chloride"), 16: (1101, "mg", "Manganese"),
        17: (1103, "µg", "Selenium"), 18: (1100, "µg", "Iodine"),
    },
    "1.5 Vitamins": {
        7: (1106, "µg", "Retinol"), 10: (1110, "µg", "Vitamin D"),
        11: (1109, "mg", "Vitamin E"), 12: (1185, "µg", "Vitamin K1 (K1 only)"),
        13: (1165, "mg", "Thiamin"), 14: (1166, "mg", "Riboflavin"),
        15: (1167, "mg", "Niacin (preformed)"), 18: (1175, "mg", "Vitamin B6"),
        19: (1178, "µg", "Vitamin B12"), 20: (1177, "µg", "Folate"),
        21: (1170, "mg", "Pantothenate"), 22: (1176, "µg", "Biotin"),
    },
    "1.12 (PUFA per 100gFood)": {
        14: (1269, "g", "cis n-6 C18:2 LA"), 16: (1270, "g", "cis n-3 C18:3 ALA"),
        26: (1271, "g", "cis n-6 C20:4 ARA"), 28: (1278, "g", "cis n-3 C20:5 EPA"),
        38: (1280, "g", "cis n-3 C22:5 DPA"), 40: (1272, "g", "cis n-3 C22:6 DHA"),
    },
}


def _parse(raw) -> Optional[tuple[float, str]]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw), Q_COMPILED
    text = str(raw).strip()
    if text in ("", "N", "n"):
        return None
    if text.lower() == "tr":
        return 0.0, Q_TRACE
    try:
        return float(text), Q_COMPILED
    except ValueError:
        return None


@lru_cache(maxsize=1)
def _sheets() -> dict[str, dict[str, tuple]]:
    from openpyxl import load_workbook

    wb = load_workbook(_XLSX, read_only=True, data_only=True)
    out: dict[str, dict[str, tuple]] = {}
    for sheet in SHEET_MAP:
        ws = wb[sheet]
        rows = {}
        for r in ws.iter_rows(values_only=True, min_row=4):
            code = str(r[0] or "").strip()
            if code:
                rows[code] = tuple(r)
        out[sheet] = rows
    wb.close()
    return out


def extract(key: str, note: str = "") -> list[SourceValue]:
    sheets = _sheets()
    out: list[SourceValue] = []
    found = False
    food = str(key)
    for sheet, cols in SHEET_MAP.items():
        row = sheets[sheet].get(str(key))
        if row is None:
            continue
        found = True
        food = f"{key} {row[1]}"
        refs = " ".join(str(row[5] or "").split())[:110]
        for i, (nid, unit, label) in cols.items():
            parsed = _parse(row[i])
            if parsed is None:
                continue
            value, quality = parsed
            out.append(SourceValue(
                source=LABEL, source_food=food, nutrient_id=nid,
                value=to_fediaf(nid, value, unit), quality=quality,
                note=f"{label}; refs: {refs}",
            ))
    if not found:
        raise KeyError(f"CoFID food code {key!r} not found")
    return out


def search(query: str) -> list[tuple[str, str]]:
    q = query.lower()
    return [
        (code, str(r[1])) for code, r in sorted(_sheets()["1.3 Proximates"].items())
        if q in str(r[1] or "").lower()
    ]
