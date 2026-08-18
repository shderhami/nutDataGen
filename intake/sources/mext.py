"""Japan MEXT Standard Tables 2015 adapter (3 volumes, EN edition).

Main volume sheet `Table`: English names row 6, INFOODS tagnames row 7, units
row 8 (1-based); data below. AA and FA volumes: use `Table 1(per 100 g EP)`
ONLY — the other tables are per g N / per 100 g total FA and must never be
read as per-100 g-food (sweep trap). Cell grammar: `-` not measured, `Tr`
trace, `(x)` estimated, plain number measured.

Form quirks encoded here:
- thiamin is reported as thiamin hydrochloride (THIAHCL): x0.887 to thiamin;
- vitamin K is the menaquinone-inclusive total (the comparator for SR's
  K1-only rows);
- MEXT is fully Japanese analytical -> quality `analysed` (or `estimated`
  for parenthesized values), no per-value n.
"""
from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path
from typing import Optional

from intake.model import (
    FORM_TOTAL_K,
    Q_ANALYSED,
    Q_COMPUTED,
    Q_ESTIMATED,
    Q_TRACE,
    SourceValue,
)
from intake.units import parse_per100g_unit, to_fediaf

LABEL = "MEXT"
_DIR = Path(__file__).resolve().parents[2] / "data" / "mext_jp"
_MAIN = _DIR / "MEXT_2015_StandardTables_Main_EN.xlsx"
_AA = _DIR / "MEXT_2015_AminoAcids_EN.xlsx"
_FA = _DIR / "MEXT_2015_FattyAcids_EN.xlsx"

THIAMIN_FROM_HCL = 0.887  # MW ratio, thiamin cation / thiamin hydrochloride

# Main volume, by tagname (row 7). Thiamin handled separately (form factor).
_MAIN_TAG_MAP: dict[str, int] = {
    "ENERC_KCAL": 1008, "WATER": 1051, "ASH": 1007, "FIBTG": 1079,
    "NA": 1093, "K": 1092, "CA": 1087, "MG": 1090, "P": 1091, "FE": 1089,
    "ZN": 1095, "CU": 1098, "MN": 1101, "ID": 1100, "SE": 1103,
    "RETOL": 1106, "VITD": 1110, "TOCPHA": 1109, "VITK": 1185,
    "RIBF": 1166, "NIA": 1167, "VITB6A": 1175, "VITB12": 1178,
    "FOL": 1177, "PANTAC": 1170, "BIOT": 1176,
}
# Main volume, by English header prefix (columns whose tagname is `－`).
_MAIN_NAME_MAP: dict[str, int] = {
    "Protein, calculated from reference": 1003,
    "Lipid": 1004,
    "Carbohydrate, total": 1005,
}
_AA_TAG_MAP: dict[str, int] = {
    "ILE": 1212, "LEU": 1213, "LYS": 1214, "MET": 1215, "CYS": 1216,
    "PHE": 1217, "TYR": 1218, "THR": 1211, "TRP": 1210, "VAL": 1219,
    "HIS": 1221, "ARG": 1220,
}
_FA_TAG_MAP: dict[str, int] = {
    "FAPU": 1293, "F18D2N6": 1269, "F18D3N3": 1270, "F20D4N6": 1271,
    "F20D5N3": 1278, "F22D5N3": 1280, "F22D6N3": 1272,
}

_PAREN_RE = re.compile(r"^\((.+)\)$")


def _parse(raw) -> Optional[tuple[float, str]]:
    """(value, quality) or None for not-measured."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw), Q_ANALYSED
    text = str(raw).strip()
    if text in ("-", "－", ""):
        return None
    if text.lower() == "tr":
        return 0.0, Q_TRACE
    m = _PAREN_RE.match(text)
    if m:
        inner = m.group(1).strip()
        if inner.lower() == "tr":
            return 0.0, Q_TRACE
        try:
            return float(inner), Q_ESTIMATED
        except ValueError:
            return None
    try:
        return float(text), Q_ANALYSED
    except ValueError:
        return None


@lru_cache(maxsize=None)  # bounded by construction: a handful of volume sheets
def _volume(path: Path, sheet: str) -> tuple[tuple, tuple, tuple, dict[str, tuple]]:
    """(names, tags, units, {item_no: row}) for one volume sheet."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    names, tags, units = rows[5], rows[6], rows[7]
    items = {str(r[1]).strip(): r for r in rows[8:] if r[1] is not None}
    return names, tags, units, items


def _norm_tag(tag) -> str:
    return str(tag or "").replace("\n", "").strip()


def _tag_columns(names: tuple, tags: tuple, tag_map: dict[str, int],
                 name_map: dict[str, int]) -> dict[int, tuple[int, str]]:
    """{column index: (nutrient_id, source column label)}."""
    cols: dict[int, tuple[int, str]] = {}
    for i, tag in enumerate(tags):
        t = _norm_tag(tag)
        if t in tag_map:
            cols[i] = (tag_map[t], t)
    for i, name in enumerate(names):
        n = " ".join(str(name or "").split())
        for prefix, nid in name_map.items():
            if n.startswith(prefix) and i not in cols:
                cols[i] = (nid, prefix)
    return cols


def _extract_volume(path: Path, key: str, tag_map: dict[str, int],
                    name_map: dict[str, int], volume_note: str) -> list[SourceValue]:
    names, tags, units, items = _volume(path, "Table" if path == _MAIN else "Table 1(per 100 g EP)")
    row = items.get(str(key))
    if row is None:
        return []
    food = f"{key} {str(row[3])[:70]}"
    out: list[SourceValue] = []
    for i, (nid, col_label) in _tag_columns(names, tags, tag_map, name_map).items():
        parsed = _parse(row[i])
        if parsed is None:
            continue
        value, quality = parsed
        unit = parse_per100g_unit(str(units[i]).replace("ｇ", "g"))
        note = f"{col_label} ({volume_note})"
        if col_label == "VITK":
            note += "; menaquinone-inclusive total K"
        if nid == 1008:
            quality = Q_ESTIMATED if quality == Q_ESTIMATED else Q_COMPUTED
            note += "; energy is always calculated"
        out.append(SourceValue(
            source=LABEL, source_food=food, nutrient_id=nid,
            value=to_fediaf(nid, value, unit), quality=quality, note=note,
            form=FORM_TOTAL_K if col_label == "VITK" else "",
        ))
    return out


def _extract_thiamin(key: str) -> list[SourceValue]:
    names, tags, units, items = _volume(_MAIN, "Table")
    row = items.get(str(key))
    if row is None:
        return []
    for i, (nid, _label) in _tag_columns(names, tags, {"THIAHCL": 1165}, {}).items():
        parsed = _parse(row[i])
        if parsed is None:
            return []
        value, quality = parsed
        return [SourceValue(
            source=LABEL, source_food=f"{key} {str(row[3])[:70]}",
            nutrient_id=nid,
            value=to_fediaf(nid, value * THIAMIN_FROM_HCL,
                            parse_per100g_unit(str(units[i]))),
            quality=quality,
            note=f"THIAHCL {value:g} x{THIAMIN_FROM_HCL} (HCl->thiamin base)",
        )]
    return []


def extract(key: str, note: str = "") -> list[SourceValue]:
    out = _extract_volume(_MAIN, key, _MAIN_TAG_MAP, _MAIN_NAME_MAP, "main volume")
    out += _extract_thiamin(key)
    out += _extract_volume(_AA, key, _AA_TAG_MAP, {}, "AA volume, per 100 g EP")
    out += _extract_volume(_FA, key, _FA_TAG_MAP, {}, "FA volume, per 100 g EP")
    return out


def search(query: str) -> list[tuple[str, str]]:
    _, _, _, items = _volume(_MAIN, "Table")
    q = query.lower()
    return [
        (no, str(r[3])[:90]) for no, r in items.items()
        if q in str(r[3] or "").lower()
    ]
