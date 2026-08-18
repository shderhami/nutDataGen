"""German BLS 4.0 (2025, MRI) adapter.

One data sheet, 7,140 foods x 145 components; per component three columns:
value `[unit/100g]`, `Datenherkunft` (origin), `Referenz`. Only the origin
column separates analytical from adopted/calculated values — trust `Analyse`
rows, flag the rest (sweep rule). Watch the units: BLS stores B6, Cu and Mn
in µg where the platform stores mg (rescaled here via the header unit).
"""
from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path
from typing import Optional

from intake.model import (
    FORM_TOTAL_K,
    Q_ANALYSED,
    Q_BORROWED,
    Q_COMPUTED,
    Q_ESTIMATED,
    Q_TRACE,
    Q_UNKNOWN,
    SourceValue,
)
from intake.units import to_fediaf

LABEL = "BLS"
_XLSX = (Path(__file__).resolve().parents[2]
         / "data" / "bls_de" / "BLS_4_0_2025_DE" / "BLS_4_0_Daten_2025_DE.xlsx")

CODE_MAP: dict[str, int] = {
    "ENERCC": 1008, "PROT625": 1003, "FAT": 1004, "WATER": 1051, "ASH": 1007,
    "FIBT": 1079, "CHO": 1005,
    "NA": 1093, "K": 1092, "CA": 1087, "MG": 1090, "P": 1091, "FE": 1089,
    "ZN": 1095, "CU": 1098, "MN": 1101, "ID": 1100, "CLD": 1088,
    "RETOL": 1106, "VITD": 1110, "TOCPHA": 1109, "VITK": 1185,
    "THIA": 1165, "RIBF": 1166, "NIA": 1167, "PANTAC": 1170, "VITB6": 1175,
    "BIOT": 1176, "FOLFD": 1177, "VITB12": 1178,
    "F18:2CN6": 1269, "F18:3CN3": 1270, "F20:4CN6": 1271,
    "F20:5CN3": 1278, "F22:5CN3": 1280, "F22:6CN3": 1272, "FAPU": 1293,
    "ILE": 1212, "LEU": 1213, "LYS": 1214, "MET": 1215, "CYSTE": 1216,
    "PHE": 1217, "TYR": 1218, "THR": 1211, "TRP": 1210, "VAL": 1219,
    "ARG": 1220, "HIS": 1221,
}
_INFO_CODES = ("VITK1", "VITK2")  # folded into the vitamin K note

_HDR_RE = re.compile(r"^(\S+) .*\[([^\]/]+)/100\s*g\]$")


def _parse(raw) -> Optional[float]:
    """BLS cells are numeric, or '-' for not-applicable/missing."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if text in ("-", ""):
        return None
    return float(text.replace(",", "."))


def _origin_quality(origin: Optional[str]) -> str:
    text = str(origin or "").lower()
    if "analyse" in text:
        return Q_ANALYSED
    if "spuren" in text:
        # trace: detection-limit information, not a measurement
        return Q_TRACE
    if ("logische" in text or "berechn" in text or "kalkul" in text
            or "rezept" in text or "reskalierung" in text or "formel" in text):
        # Logische Null/Annahme are definitional zeros/assumptions, never
        # measurements (audit 2026-08-18: an unclassified Logische Null was
        # counted as independent confirming evidence).
        return Q_COMPUTED
    if "schätz" in text:
        return Q_ESTIMATED
    if any(m in text for m in ("übernommen", "literatur", "datenbank",
                               "aggregation", "labelangabe")):
        # Aggregation rows name the pooled databases in Referenz (frida/NO/USDA
        # in practice); Labelangabe is a manufacturer's package declaration —
        # neither is independent German analytical work.
        return Q_BORROWED
    return Q_UNKNOWN


@lru_cache(maxsize=1)
def _header() -> dict[str, tuple[int, str]]:
    """{code: (value column, unit)} from the header row."""
    from openpyxl import load_workbook

    wb = load_workbook(_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = next(ws.iter_rows(values_only=True, max_row=1))
    wb.close()
    cols: dict[str, tuple[int, str]] = {}
    for i, h in enumerate(hdr):
        m = _HDR_RE.match(str(h or ""))
        if m:
            cols[m.group(1)] = (i, m.group(2).strip())
    return cols


@lru_cache(maxsize=32)
def _row(code: str) -> Optional[tuple]:
    from openpyxl import load_workbook

    wb = load_workbook(_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hit = None
    for r in ws.iter_rows(values_only=True, min_row=2):
        if str(r[0] or "") == code:
            hit = r
            break
    wb.close()
    return hit


def extract(key: str, note: str = "") -> list[SourceValue]:
    cols = _header()
    row = _row(str(key))
    if row is None:
        raise KeyError(f"BLS code {key!r} not found")
    food = f"{key} {row[2]}"
    out: list[SourceValue] = []
    k_forms = {}
    for code in _INFO_CODES:
        ci, _unit = cols[code]
        v = _parse(row[ci])
        if v is not None:
            k_forms[code] = v
    for code, nid in CODE_MAP.items():
        ci, unit = cols[code]
        val = _parse(row[ci])
        if val is None:
            continue
        origin = row[ci + 1]
        ref = row[ci + 2]
        extra = ""
        if code == "VITK" and k_forms:
            extra = "; " + ", ".join(f"{c}={v:g}µg" for c, v in k_forms.items())
        out.append(SourceValue(
            source=LABEL, source_food=food, nutrient_id=nid,
            value=to_fediaf(nid, val, unit),
            quality=_origin_quality(origin),
            note=f"{code} [{origin}] {str(ref or '')[:80]}{extra}".strip(),
            form=FORM_TOTAL_K if code == "VITK" else "",
        ))
    return out


def search(query: str) -> list[tuple[str, str]]:
    from openpyxl import load_workbook

    q = query.lower()
    wb = load_workbook(_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hits = [
        (str(r[0]), str(r[2]))
        for r in ws.iter_rows(values_only=True, min_row=2)
        if q in str(r[2] or "").lower() or q in str(r[1] or "").lower()
    ]
    wb.close()
    return hits
