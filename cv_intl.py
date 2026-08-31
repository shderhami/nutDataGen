#!/usr/bin/env python3
"""International CV observations for cv-v8 pooling, and the sigma^2 calibration.

An "observation" is a foreign table's spread measurement (n, min, max) for one of
our (food_id, nutrient_id) cells, matched by food. Pooling credits it at the
DISCOUNTED effective n' = 1 / (1/n + 2*sigma^2), where sigma^2 is the measured
between-population variance of ln(CV) — see cv_config.INTL_CV_SIGMA2 for the
derivation record. Traceability: every observation carries its source table,
source food entry, and citation; resolve_cv records the full pooling
decomposition in cv_method_inputs.

Two entry points:
  build   — regenerate data/cv_curation/intl_cv_observations.csv from the pinned
            FCDB workbook + the curated FOOD_MAP/EXCLUDES below (deterministic).
  sigma   — recompute sigma^2 from the observations file joined against USDA bulk
            stats; compare with the applied cv_config.INTL_CV_SIGMA2.

Curation lives in code as data (FOOD_MAP + EXCLUDES + eligibility rules), so the
CSV is reproducible and the whole chain is covered by dataset/config hashing.
"""
from __future__ import annotations

import csv
import math
import statistics
from pathlib import Path
from typing import Optional

import cv_config
import cv_stats

INTL_CSV = Path(__file__).parent / "data" / "cv_curation" / "intl_cv_observations.csv"
_FCDB_XLSX = Path(__file__).parent / "data" / "fcdb_dk" / "FCDB_6.1_Dataset.xlsx"

# our food_id -> (FCDB FoodID, FCDB food name as recorded)
# 10048 (skinless thigh, 2026-08-18) and 10054 (skinless breast, 2026-08-26)
# are deliberately UNMAPPED: FCDB has no cut-specific skinless entries; 795
# "Chicken, flesh only" is a whole-bird blend whose dispersion describes
# neither cut's population.
# 2026-08-29 (v8.3): FCDB 1019 chuck REASSIGNED 10035 -> 10055. build() maps one
# FCDB food to one of ours, and the Danish chuck frame (fat 13.1, n=6 range
# 9.2-15.4) matches the trimmed chuck roll (13.9), not the untrimmed blade
# roast (17.8) it was loosely paired with - that pairing needed fat/water/ash
# EXCLUDES that the honest frame match makes unnecessary. 10035 is now
# UNMAPPED (its FND n=8 stats still carry its measured tiers).
FOOD_MAP = {
    10050: (641, "Heart, beef, raw"),
    10055: (1019, "Beef, chuck, raw"),
    10056: (56, "Squash, raw"),  # Danish "squash" = courgette/zucchini (2020 veg study, n=8 on 19 params)
    10057: (764, "Beans, green, raw"),  # older Danish composition surveys (n=4-13); B2 range corrupt in workbook (min>max, self-drops)
    10002: (712, "Liver, broiler or fryer, raw"),
    10013: (742, "Liver, ox, raw"),
    10021: (940, 'Beef, topside "cap on", raw'),
    10038: (789, "Pork, hand, lean, raw"),
    10037: (739, "Pork, loin, lean, raw"),
    10014: (1658, "Salmon, atlantic, aquaculture, raw"),
    10009: (387, "Mussel, raw"),
    10019: (1745, "Butternut squash, raw"),
    10010: (1662, "Eggs, chicken, battery hens, raw"),
    10031: (1230, "Egg, chicken, yolk, raw"),
    10018: (215, "Eggs, chicken, egg white, raw"),
}

# FCDB ParameterName -> our nutrient_id (definition-comparable only)
PARAM_MAP = {
    "Protein": 1003, "Fat": 1004, "Ash": 1007, "Water": 1051,
    "Retinol": 1106, "alpha-Tocopherol": 1109, "Vitamin D3": 1110,
    "Thiamin (Vitamin B1)": 1165, "Riboflavin (Vitamin B2)": 1166, "Niacin": 1167,
    "Vitamin B6": 1175, "Vitamin B12": 1178, "Pantothenic acid": 1170,
    "Folate": 1177, "Biotin": 1176,
    "Sodium": 1093, "Potassium": 1092, "Calcium": 1087, "Magnesium": 1090,
    "Iron": 1089, "Copper": 1098, "Zinc": 1095, "Manganese": 1101,
    "Selenium": 1103, "Phosphorus": 1091, "Iodine": 1100, "Chloride": 1088,
}

# (food_id, nutrient_id) pairs excluded by curation, with the reason kept here
# as the audit record. Frame mismatches make the foreign CV describe something
# other than our cell's population spread.
EXCLUDES = {
    # (10035, 1004/1051/1007) trim-mismatch excludes retired 2026-08-29: FCDB
    # 1019 now maps to the frame-matched 10055, where fat/water/ash are honest.
    (10056, 1176): "cell's own literature_range stats ARE this FCDB row (study 2127) — double-count guard",
    (10057, 1098): "1977-80 contaminant-monitoring survey (src 1082), not retail population spread — pork Cu/Zn precedent",
    (10057, 1095): "1977-80 contaminant-monitoring survey (src 1082) — pork Cu/Zn precedent",
    (10021, 1004): "trim mismatch: cap-on topside vs 1/8-inch trim",
    (10021, 1003): "trim mismatch",
    (10038, 1098): "FCDB multi-decade contaminant monitoring series (n=266)",
    (10038, 1095): "FCDB multi-decade contaminant monitoring series (n=266)",
    (10037, 1098): "FCDB multi-decade contaminant monitoring series (n=266)",
    (10037, 1095): "FCDB multi-decade contaminant monitoring series (n=266)",
}

# our food_id -> FDC ids whose bulk stats form the US side of calibration pairs
FDC_MAP = {
    10050: (168625,),
    10055: (170814,),
    10056: (169291, 2685568),
    10057: (169961, 2346400),
    10002: (171060,), 10013: (169451,), 10021: (173997, 746761),
    10038: (168260,), 10037: (168314, 2646168),
    10014: (175167, 2684441), 10009: (174216,), 10019: (169295, 2685570),
    10010: (171287, 748967), 10031: (172184, 748236), 10018: (172183, 747997),
}

MIN_N = 4          # foreign observations below this add ~nothing and skew wan-range
CV_CAP = 1.2       # censoring/detection-limit artifacts live above this


def _f(x) -> Optional[float]:
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def build() -> int:
    """Regenerate the observations CSV from the pinned FCDB workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(_FCDB_XLSX, read_only=True, data_only=True)
    ws = wb["Data_Normalised"]
    it = ws.iter_rows(values_only=True)
    header = [str(c) for c in next(it)]
    idx = {h: i for i, h in enumerate(header)}
    by_fcdb = {fcdb_id: (food_id, name) for food_id, (fcdb_id, name) in FOOD_MAP.items()}
    rows = []
    for r in it:
        hit = by_fcdb.get(r[idx["FoodID"]])
        if hit is None:
            continue
        food_id, src_food = hit
        nid = PARAM_MAP.get(r[idx["ParameterName"]])
        if nid is None or (food_id, nid) in EXCLUDES:
            continue
        val, mn, mx = _f(r[idx["ResVal"]]), _f(r[idx["Min"]]), _f(r[idx["Max"]])
        n = r[idx["NumberOfDeterminations"]]
        n = int(n) if str(n).isdigit() else None
        if val is None or mn is None or mx is None or n is None or n < MIN_N:
            continue
        if mn <= 0:  # detection-limit censoring: range no longer reflects spread
            continue
        cv = cv_stats.cv_from_range(val, mn, mx, n)
        if cv is None or cv > CV_CAP:
            continue
        rows.append({
            "food_id": food_id, "nutrient_id": nid,
            "source": "FCDB 6.1", "source_food_id": r[idx["FoodID"]],
            "source_food": src_food, "param": r[idx["ParameterName"]],
            "value": val, "n": n, "min": mn, "max": mx, "cv_wan": round(cv, 6),
            "source_refs": str(r[idx["Source"]]),
        })
    wb.close()
    rows.sort(key=lambda x: (x["food_id"], x["nutrient_id"]))
    INTL_CSV.parent.mkdir(exist_ok=True)
    with open(INTL_CSV, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    return len(rows)


def read_observations() -> dict:
    """(food_id, nutrient_id) -> {cv, n, label} for resolve_cv pooling."""
    out = {}
    if not INTL_CSV.exists():
        return out
    with open(INTL_CSV) as fh:
        for r in csv.DictReader(fh):
            key = (int(r["food_id"]), int(r["nutrient_id"]))
            out[key] = {
                "cv": float(r["cv_wan"]), "n": int(r["n"]),
                "label": f'{r["source"]} {r["source_food_id"]} "{r["source_food"]}" {r["param"]}',
            }
    return out


def effective_n(n: int, sigma2: float) -> float:
    """Discounted worth of a foreign observation, in same-population samples."""
    return 1.0 / (1.0 / n + 2.0 * sigma2)


def recompute_sigma2() -> Optional[dict]:
    """DerSimonian-Laird-style sigma^2 from observations paired with US bulk stats."""
    us: dict = {}
    for path in (cv_config.FDC_SRL_DIR / "food_nutrient.csv",
                 cv_config.FDC_FDN_DIR / "food_nutrient.csv"):
        with open(path) as fh:
            for row in csv.reader(fh):
                if row[0] == "id":
                    continue
                nid = int(row[2])
                amount, dp, mn, mx = _f(row[3]), int(row[4] or 0), _f(row[6]), _f(row[7])
                if dp >= MIN_N and amount and mn is not None and mx is not None:
                    us[(row[1], nid)] = (amount, dp, mn, mx)
    # our food_id -> the fdc ids we know (from the ingredients table would be
    # cleaner, but keeping this file DB-independent: derive from bulk by trying
    # both ids recorded in the observations pairing below).
    lrs, noises = [], []
    for r in csv.DictReader(open(INTL_CSV)):
        food_id, nid = int(r["food_id"]), int(r["nutrient_id"])
        for fdc in FDC_MAP.get(food_id, ()):
            u = us.get((str(fdc), nid))
            if not u:
                continue
            us_cv = cv_stats.cv_from_range(u[0], u[2], u[3], u[1])
            dk_cv = float(r["cv_wan"])
            if us_cv is None or not (0.05 <= us_cv <= CV_CAP and 0.05 <= dk_cv <= CV_CAP):
                continue
            lrs.append(math.log(us_cv / dk_cv))
            noises.append(1 / (2 * u[1]) + 1 / (2 * int(r["n"])))
            break
    if len(lrs) < 10:
        return None
    obs = statistics.variance(lrs)
    return {"pairs": len(lrs), "bias": statistics.mean(lrs),
            "observed_var": obs, "noise_var": statistics.mean(noises),
            "sigma2": max(0.0, obs - statistics.mean(noises))}


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "build":
        print(f"wrote {build()} observations -> {INTL_CSV}")
    else:
        res = recompute_sigma2()
        if res is None:
            print("not enough pairs to calibrate")
        else:
            print(f"pairs={res['pairs']} bias={res['bias']:+.3f} "
                  f"sigma2(recomputed)={res['sigma2']:.3f} "
                  f"sigma2(applied)={cv_config.INTL_CV_SIGMA2:.3f}")
